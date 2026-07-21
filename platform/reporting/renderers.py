"""Report renderers — one function per output format.

All four take the same inputs (a ReportSpec plus a branding dict) and return
bytes. Adding a format = one function + one FORMATS entry in services.py.

Branding dict keys: company (str), generated_by (str), generated_at (str).
"""

from __future__ import annotations

import csv
import io

from django.template.loader import render_to_string

from reporting.spec import ReportSpec

_ALIGN_XLSX = {"left": "left", "right": "right", "center": "center"}


def render_csv(spec: ReportSpec, branding: dict[str, str]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([column.label for column in spec.columns])
    for row in spec.rows:
        writer.writerow([spec.cell(row, column) for column in spec.columns])
    return buffer.getvalue().encode("utf-8")


def render_xlsx(spec: ReportSpec, branding: dict[str, str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.title[:31] or "Report"

    sheet.append([spec.title])
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.append([branding["company"]])
    sheet.append([f"Generated {branding['generated_at']} by {branding['generated_by']}"])
    if spec.filters:
        sheet.append(["Filters: " + ", ".join(f"{k}={v}" for k, v in spec.filters.items())])
    sheet.append([])

    header_row = sheet.max_row + 1
    sheet.append([column.label for column in spec.columns])
    for index in range(1, len(spec.columns) + 1):
        sheet.cell(row=header_row, column=index).font = Font(bold=True)

    for row in spec.rows:
        sheet.append([spec.cell(row, column) for column in spec.columns])

    for index, column in enumerate(spec.columns, start=1):
        width = (
            max(len(column.label), *(len(spec.cell(row, column)) for row in spec.rows), 8)
            if spec.rows
            else len(column.label) + 2
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 60)
        alignment = Alignment(horizontal=_ALIGN_XLSX.get(column.align, "left"))
        for row_index in range(header_row, sheet.max_row + 1):
            sheet.cell(row=row_index, column=index).alignment = alignment

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_html(spec: ReportSpec, branding: dict[str, str]) -> bytes:
    rendered_rows = [
        [{"value": spec.cell(row, column), "align": column.align} for column in spec.columns]
        for row in spec.rows
    ]
    html = render_to_string(
        "reporting/report.html",
        {"spec": spec, "branding": branding, "rendered_rows": rendered_rows},
    )
    return html.encode("utf-8")


def render_pdf(spec: ReportSpec, branding: dict[str, str]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    styles = getSampleStyleSheet()

    def _decorate(canvas, _doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawString(15 * mm, 10 * mm, f"{branding['company']} — {spec.title}")
        canvas.drawRightString(195 * mm, 10 * mm, f"Page {canvas.getPageNumber()}")
        if spec.watermark:
            canvas.setFont("Helvetica-Bold", 60)
            canvas.setFillColorRGB(0.88, 0.88, 0.88)
            canvas.translate(105 * mm, 148 * mm)
            canvas.rotate(45)
            canvas.drawCentredString(0, 0, spec.watermark)
        canvas.restoreState()

    story = [
        Paragraph(spec.title, styles["Title"]),
        Paragraph(branding["company"], styles["Normal"]),
        Paragraph(
            f"Generated {branding['generated_at']} by {branding['generated_by']}",
            styles["Normal"],
        ),
    ]
    if spec.subtitle:
        story.append(Paragraph(spec.subtitle, styles["Normal"]))
    if spec.filters:
        filters = ", ".join(f"{k}={v}" for k, v in spec.filters.items())
        story.append(Paragraph(f"Filters: {filters}", styles["Normal"]))
    story.append(Spacer(1, 6 * mm))

    data = [[column.label for column in spec.columns]] + [
        [spec.cell(row, column) for column in spec.columns] for row in spec.rows
    ]
    table = Table(data, repeatRows=1)
    align_commands = [
        ("ALIGN", (index, 0), (index, -1), column.align.upper())
        for index, column in enumerate(spec.columns)
        if column.align in ("right", "center")
    ]
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                *align_commands,
            ]
        )
    )
    story.append(table)

    SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=spec.title,
    ).build(story, onFirstPage=_decorate, onLaterPages=_decorate)
    return buffer.getvalue()
