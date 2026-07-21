"""Reporting engine tests — renderers, export pipeline, history API."""

from __future__ import annotations

import csv
import io

import pytest
from audit.models import AuditLog
from reporting.services import ReportService
from reporting.spec import ReportColumn, ReportSpec
from shared.events import Events
from shared.exceptions import ValidationError
from storage.services import StorageService


@pytest.fixture(autouse=True)
def _storage_root(settings, tmp_path):
    settings.STORAGE_LOCAL_PATH = str(tmp_path / "objects")


@pytest.fixture
def owner(make_user, tenant, owner_role, grant):
    user = make_user(email="owner@acme.test", username="owner", tenant=tenant)
    grant(user, owner_role)
    return user


@pytest.fixture
def spec():
    return ReportSpec(
        key="test-bills",
        title="Purchase bills",
        module="purchase",
        columns=[
            ReportColumn("seller", "Seller"),
            ReportColumn("total", "Total", align="right"),
        ],
        rows=[
            {"seller": "Vendor Inc", "total": "150.00"},
            {"seller": "Steel & Co", "total": "99.50"},
        ],
        filters={"status": "confirmed"},
        watermark="DRAFT",
    )


def test_csv_renderer(tenant, spec):
    payload, content_type, filename = ReportService().render(spec, "csv", tenant=tenant)

    rows = list(csv.reader(io.StringIO(payload.decode("utf-8"))))
    assert rows[0] == ["Seller", "Total"]
    assert rows[1] == ["Vendor Inc", "150.00"]
    assert content_type == "text/csv"
    assert filename.startswith("test-bills-") and filename.endswith(".csv")


def test_xlsx_renderer_opens_and_contains_data(tenant, spec):
    from openpyxl import load_workbook

    payload, content_type, _ = ReportService().render(spec, "xlsx", tenant=tenant)
    sheet = load_workbook(io.BytesIO(payload)).active

    values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    flat = [v for row in values for v in row if v is not None]
    assert "Purchase bills" in flat
    assert tenant.name in flat
    assert "Vendor Inc" in flat
    assert content_type.endswith("spreadsheetml.sheet")


def test_html_renderer_brands_and_watermarks(tenant, spec, owner):
    payload, content_type, _ = ReportService().render(spec, "html", tenant=tenant, actor=owner)
    html = payload.decode("utf-8")

    assert "Purchase bills" in html
    assert tenant.name in html
    assert "Steel &amp; Co" in html
    assert "DRAFT" in html
    assert "status=confirmed" in html
    assert owner.email in html
    assert content_type == "text/html"


def test_pdf_renderer_produces_pdf(tenant, spec):
    payload, content_type, _ = ReportService().render(spec, "pdf", tenant=tenant)
    assert payload.startswith(b"%PDF")
    assert len(payload) > 1000
    assert content_type == "application/pdf"


def test_unknown_format_rejected(tenant, spec):
    with pytest.raises(ValidationError):
        ReportService().render(spec, "docx", tenant=tenant)


def test_export_stores_file_and_records_history(tenant, spec, owner):
    export = ReportService().export(spec, "csv", tenant=tenant, actor=owner)

    assert export.row_count == 2
    assert export.module == "purchase"
    assert export.file is not None
    assert export.file.category == "report"
    assert StorageService().open(export.file).startswith(b"Seller,Total")
    assert AuditLog.objects.filter(action=Events.REPORT_EXPORTED, module="reporting").exists()


def test_history_api_scoped_and_permissioned(
    tenant, other_tenant, spec, owner, make_user, grant, owner_role, auth_client
):
    export = ReportService().export(spec, "pdf", tenant=tenant, actor=owner)
    client = auth_client(owner)

    listed = client.get("/api/v1/reporting/exports/")
    assert [row["id"] for row in listed.data["data"]] == [str(export.id)]

    url_response = client.get(f"/api/v1/reporting/exports/{export.id}/url/")
    assert url_response.status_code == 200
    assert "token=" in url_response.data["url"]

    outsider = make_user(email="owner@globex.test", username="globex", tenant=other_tenant)
    grant(outsider, owner_role)
    assert auth_client(outsider).get("/api/v1/reporting/exports/").data["data"] == []

    nobody = make_user(email="nobody@acme.test", username="nobody", tenant=tenant)
    assert auth_client(nobody).get("/api/v1/reporting/exports/").status_code == 403
