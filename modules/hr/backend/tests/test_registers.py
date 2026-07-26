"""Statutory register generation.

Exercises the real migrated generator against the real company template — not a
mock — so the openpyxl code that fills the ten official forms is genuinely
covered. Then asserts the compliance guarantees the legacy app made: the
template is never written to, versions accumulate, the archive is hash-verified,
and generating locks the month.
"""

from __future__ import annotations

from datetime import date

import pytest
from openpyxl import load_workbook
from periods.services import PeriodService

from hr.backend.models import GenerationLog
from hr.backend.services.registers.config import get_settings

pytestmark = pytest.mark.django_db

ATTENDANCE_URL = "/api/v1/hr/attendance/"
PAYROLL_RUN_URL = "/api/v1/hr/payroll/run/"
REGISTERS_URL = "/api/v1/hr/registers/"
GENERATE_URL = "/api/v1/hr/registers/generate/"

YEAR, MONTH = 2026, 4


@pytest.fixture
def month_with_payroll(auth_client, owner, employee_factory, salary_rule_factory, pay_rules):
    """One employee, a full month present, payroll run — ready to generate."""
    employee = employee_factory(doj=date(2024, 1, 1))
    salary_rule_factory(employee=employee)
    client = auth_client(owner)
    client.post(
        ATTENDANCE_URL,
        {
            "year": YEAR,
            "month": MONTH,
            "records": [
                {
                    "employee_id": str(employee.pk),
                    "days": [{"day": d, "status": "P"} for d in range(1, 27)],
                }
            ],
        },
        format="json",
    )
    client.post(PAYROLL_RUN_URL, {"year": YEAR, "month": MONTH}, format="json")
    return client, employee


class TestTemplateIsSacred:
    def test_the_template_ships_with_the_module(self):
        # It cannot be regenerated from code, so it must be committed.
        assert get_settings().template_full_path.exists()

    def test_all_ten_form_mappings_ship_with_the_module(self):
        mappings = sorted(p.name for p in get_settings().mappings_full_path.glob("*.json"))
        assert len(mappings) == 10
        assert "form_b.json" in mappings
        assert "form_xxvi.json" in mappings
        assert "in_out_register.json" in mappings

    def test_generation_never_modifies_the_template(self, month_with_payroll):
        client, _ = month_with_payroll
        template = get_settings().template_full_path
        before = template.read_bytes()

        response = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        assert response.status_code == 201, response.data
        assert template.read_bytes() == before


class TestGeneration:
    def test_generate_produces_a_readable_workbook(self, month_with_payroll):
        client, _ = month_with_payroll

        response = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        assert response.status_code == 201, response.data
        log = GenerationLog.objects.get(pk=response.data["id"])
        assert log.version == 1
        assert len(log.file_hash) == 64
        assert log.stored_file is not None

        download = client.get(f"{REGISTERS_URL}{log.pk}/download/")
        assert download.status_code == 200
        assert download["Content-Disposition"].startswith("attachment;")

        # It really is a workbook, and it really has the statutory sheets.
        import io

        workbook = load_workbook(io.BytesIO(download.content))
        assert len(workbook.sheetnames) > 1
        workbook.close()

    def test_generating_before_payroll_is_refused(self, auth_client, owner):
        response = auth_client(owner).post(
            GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json"
        )

        # No figures means no defensible register.
        assert response.status_code == 422
        assert not GenerationLog.objects.exists()

    def test_the_working_copy_is_cleaned_up(self, month_with_payroll, tmp_path, monkeypatch):
        client, _ = month_with_payroll
        # Isolate the build directory so this asserts on our own run only.
        monkeypatch.setenv("HR_REGISTERS_WORK_DIR", str(tmp_path))
        assert get_settings().generated_full_path == tmp_path

        response = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        assert response.status_code == 201, response.data
        # The durable copy lives in platform storage, not on local disk.
        assert list(tmp_path.glob("*.xlsx")) == []


class TestVersioningAndIntegrity:
    def test_regenerating_adds_a_version_and_keeps_the_old_one(self, month_with_payroll, tenant):
        client, _ = month_with_payroll
        first = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")
        # Generating locked the month; reopen it to regenerate, as an operator would.
        PeriodService().unlock(
            tenant=tenant, year=YEAR, month=MONTH, reason="Correction before submission"
        )

        second = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        assert first.status_code == 201
        assert second.status_code == 201, second.data
        versions = sorted(
            GenerationLog.objects.filter(year=YEAR, month=MONTH).values_list("version", flat=True)
        )
        assert versions == [1, 2]
        # The first submission is still downloadable — nothing was overwritten.
        assert client.get(f"{REGISTERS_URL}{first.data['id']}/download/").status_code == 200

    def test_verify_confirms_an_untampered_archive(self, month_with_payroll):
        client, _ = month_with_payroll
        created = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        response = client.get(f"{REGISTERS_URL}{created.data['id']}/verify/")

        assert response.status_code == 200, response.data
        assert response.data["valid"] is True
        assert response.data["expected_hash"] == response.data["actual_hash"]

    def test_verify_detects_a_tampered_archive(self, month_with_payroll):
        client, _ = month_with_payroll
        created = client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")
        log = GenerationLog.objects.get(pk=created.data["id"])

        # Simulate the bytes being swapped underneath the recorded digest.
        from storage.providers import get_provider

        get_provider().put(log.stored_file.key, b"tampered", log.stored_file.content_type)

        response = client.get(f"{REGISTERS_URL}{log.pk}/verify/")

        assert response.status_code == 200
        assert response.data["valid"] is False


class TestGenerationLocksTheMonth:
    def test_generating_locks_the_period(self, month_with_payroll, tenant):
        client, _ = month_with_payroll

        client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        assert PeriodService().is_locked(tenant=tenant, year=YEAR, month=MONTH) is True

    def test_attendance_is_frozen_after_generation(self, month_with_payroll):
        client, employee = month_with_payroll
        client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        response = client.post(
            ATTENDANCE_URL,
            {
                "year": YEAR,
                "month": MONTH,
                "records": [
                    {"employee_id": str(employee.pk), "days": [{"day": 27, "status": "P"}]}
                ],
            },
            format="json",
        )

        assert response.status_code == 409

    def test_history_lists_generations_newest_first(self, month_with_payroll, tenant):
        client, _ = month_with_payroll
        client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")
        PeriodService().unlock(tenant=tenant, year=YEAR, month=MONTH, reason="Re-issue")
        client.post(GENERATE_URL, {"year": YEAR, "month": MONTH}, format="json")

        response = client.get(REGISTERS_URL)

        assert response.status_code == 200
        assert [row["version"] for row in response.data["data"]] == [2, 1]
