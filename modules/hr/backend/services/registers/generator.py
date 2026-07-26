"""
Excel Generator — Orchestrates generation of all statutory sheets.

Generates fully computed values (not Excel formulas) into all 10 statutory
sheets, plus hidden source data sheets for traceability.
"""

import gc
import logging
import shutil
from calendar import monthrange
from copy import copy
from datetime import datetime, time

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from hr.backend.services.registers.config import get_settings
from hr.backend.services.registers.mapper import ExcelMapper
from hr.backend.services.registers.writer import ExcelWriter, safe_set

logger = logging.getLogger(__name__)

# Full-day equivalents used when a day is marked worked but has no swipe
# times to derive hours from (P = full shift, HD = half day).
WORKED_DEFAULT_HOURS = {"P": 8.0, "HD": 4.0}

# Status-code groups mirroring app.services.payroll_engine so the Excel
# COUNTIF formulas count the same codes the payroll engine does.
LEAVE_CODES = ("PL", "CL", "SL", "L", "CO")
HOLIDAY_CODES = ("NH", "WH")
WEEK_OFF_CODES = ("WO", "W/O")

# In-out register: font applied ONLY to cells where attendance values are
# written. Borders, merges, row heights and column widths stay untouched —
# only the font name/size on written cells changes.
ATTENDANCE_FONT_NAME = "Arial Narrow"
ATTENDANCE_FONT_SIZE = 28

# Statuses flagged RED in the register: Absent, Leave (PL/CL/SL/L), Half-Day.
# Present, Holiday (NH/WH), Weekly Off (WO), Comp Off (CO), OT and normal
# working days are never colored.
RED_STATUSES = {"A", "PL", "CL", "SL", "L", "HD"}
RED_FILL = PatternFill(
    start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"
)  # light pink-red
YELLOW_FILL = PatternFill(
    start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"
)  # light cream-yellow
BLUE_FILL = PatternFill(
    start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid"
)  # light sky-blue
NO_FILL = PatternFill(fill_type=None)  # explicit "no fill" to reset stale styles

DEFAULT_CERTIFICATION = (
    "  Certified that the above said employees timings were captured from the "
    "Gate entry register / Swipe card records maintained at the work site."
)


def _countif_sum(rng: str, codes) -> str:
    """Build =COUNTIF(rng,"X")+COUNTIF(rng,"Y")... for a set of status codes."""
    if not codes:
        return "=0"
    return "=" + "+".join(f'COUNTIF({rng},"{c}")' for c in codes)


class ExcelGenerator:
    """Orchestrates the generation of the HR statutory register workbook."""

    def __init__(self):
        self.settings = get_settings()
        self.mapper = ExcelMapper()

    def generate(self, year: int, month: int, payload_data: dict) -> str:
        """
        Generates the Excel file for a given month and year.
        Returns the filename of the generated file.
        """
        template_path = self.settings.template_full_path
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found at {template_path}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"HR_Register_{year}_{month:02d}_{timestamp}.xlsx"
        output_path = self.settings.generated_full_path / filename

        # Copy template
        shutil.copy2(template_path, output_path)
        logger.info(f"Copied template to {output_path}")

        wb = openpyxl.load_workbook(output_path)
        writer = ExcelWriter(wb, self.mapper)

        employees = payload_data.get("employees", [])
        attendance = payload_data.get("attendance", [])
        payroll = payload_data.get("payroll", [])

        month_name = datetime(year, month, 1).strftime("%b-%y")
        _, days_in_month = monthrange(year, month)

        # Prepare header data
        header_data = {
            "year": year,
            "month_date": f"{month:02d}-{year}",
            "month_name": month_name,
            "contractor_name": self.settings.CONTRACTOR_NAME,
            "contractor_address": self.settings.CONTRACTOR_ADDRESS,
            "principal_employer": self.settings.PRINCIPAL_EMPLOYER,
            "work_location": self.settings.WORK_LOCATION,
            "main_contractor": self.settings.MAIN_CONTRACTOR,
            "esic_code": self.settings.ESIC_CODE,
            "epf_code": self.settings.EPF_CODE,
        }

        # Build lookup structures
        emp_attendance = self._group_attendance_by_employee(attendance)
        payroll_by_emp = {p.employee_id: p for p in payroll}

        # ── Generate each statutory sheet ──
        self._write_form_xxvi(
            writer, header_data, employees, emp_attendance, payroll_by_emp, days_in_month
        )
        self._write_form_xxvii(writer, header_data, employees, payroll_by_emp)
        self._write_form_xxviii(writer, header_data, employees, payroll_by_emp)
        self._write_form_b(writer, header_data, employees, payroll_by_emp)
        self._write_form_c(writer, header_data)
        self._write_form_d(writer, header_data, employees, payroll_by_emp)
        self._write_form_vi(writer, header_data, employees)
        self._write_deduction_sheet(writer, header_data, employees, payroll_by_emp)
        self._write_in_out_register(writer, header_data, employees, emp_attendance, days_in_month)
        self._write_esic_epf_tracking(writer, header_data, employees)

        # ── Hidden source data sheets for traceability ──
        self._write_source_data(wb, employees, attendance, payroll)

        wb.save(output_path)
        wb.close()
        del wb  # release the in-memory workbook immediately
        gc.collect()

        logger.info(f"Generated Excel file: {filename}")
        return filename

    def _group_attendance_by_employee(self, attendance):
        """Group attendance records by employee_id → day → record."""
        emp_attendance = {}
        for att in attendance:
            if att.employee_id not in emp_attendance:
                emp_attendance[att.employee_id] = {}
            emp_attendance[att.employee_id][att.day] = {
                "in_time": att.in_time,
                "out_time": att.out_time,
                "status": att.status,
            }
        return emp_attendance

    @staticmethod
    def _parse_hhmm(value):
        """Parse an 'HH:MM' / 'HH:MM:SS' string into datetime.time, else None."""
        if not value or not isinstance(value, str):
            return None
        parts = value.strip().split(":")
        try:
            h, m = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
        except (ValueError, IndexError):
            return None
        if 0 <= h <= 23 and 0 <= m <= 59:
            return time(h, m)
        return None

    @staticmethod
    def _last_used_row(ws, start_row: int, max_col: int, scan_limit: int) -> int:
        """
        Last row containing any value within a bounded band. Some template
        sheets declare corrupt dimensions (A1:BT1048576), so ws.max_row is
        never trustworthy here — always scan a bounded range instead.
        """
        last = start_row - 1
        for r in range(start_row, scan_limit + 1):
            for c in range(1, max_col + 1):
                if ws.cell(row=r, column=c).value not in (None, ""):
                    last = r
                    break
        return last

    # ─── Form No. XXVI — Attendance Register ──────────────────────────────────

    def _write_form_xxvi(
        self, writer, header_data, employees, emp_attendance, payroll_by_emp, days_in_month
    ):
        sheet_name = "Form No.XXVI"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 7)
        ws = writer.wb[sheet_name]

        daily = mapping.get("daily_data", {})
        day_min_col = openpyxl.utils.column_index_from_string(daily.get("start_col", "I"))
        max_day_slots = daily.get("max_days", 31)
        day_max_col = day_min_col + max_day_slots - 1  # AM
        first_day = get_column_letter(day_min_col)
        last_day = get_column_letter(day_max_col)
        last_clear_col = openpyxl.utils.column_index_from_string("AW")

        # Template carries junk merges inside the daily grid (e.g. AC8:AM8)
        # that would collapse per-day writes into one anchor cell.
        writer.unmerge_data_region(sheet_name, start_row, day_min_col, day_max_col + 8)

        # Clear the stale data band (old employee rows + old totals rows) so
        # a smaller employee set never leaves ghost rows behind.
        stale_end = self._last_used_row(ws, start_row, last_clear_col, start_row + 200)
        for r in range(start_row, stale_end + 1):
            for c in range(1, last_clear_col + 1):
                cell = ws.cell(row=r, column=c)
                # Structural merges (multi-col labels) live outside the unmerged
                # day grid; their non-anchor cells are read-only. Skipping them
                # preserves the template layout — the merge's value sits in its
                # anchor (a normal cell) which is still cleared when reached.
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        for idx, emp in enumerate(employees):
            row = start_row + idx
            pr = payroll_by_emp.get(emp.id)

            writer.write_row_data(
                sheet_name,
                row,
                {
                    "serial_no": idx + 1,
                    "employee_code": emp.employee_code,
                    "employee_name": emp.employee_name,
                    "age_gender": emp.age_gender_display,
                    "address": emp.address or "",
                    "designation": emp.designation,
                    "father_husband_name": emp.father_husband_name or "",
                    "date_of_joining": emp.date_of_joining.strftime("%d-%m-%Y")
                    if emp.date_of_joining
                    else "",
                    # Not derivable from status codes — taken from payroll:
                    "holiday_working_days": pr.holiday_working_days if pr else 0,
                    "termination_date": emp.date_of_leaving.strftime("%d-%m-%Y")
                    if emp.date_of_leaving
                    else "",
                },
            )

            # Daily grid holds HOURS for worked days (from swipe times when
            # available, else the P/HD full-day equivalent) and the status
            # code text otherwise, so =SUM over the row is meaningful.
            daily_records = emp_attendance.get(emp.id, {})
            for day_num in range(1, days_in_month + 1):
                rec = daily_records.get(day_num)
                if not rec:
                    continue
                cell = ws.cell(row=row, column=day_min_col + day_num - 1)
                t_in = self._parse_hhmm(rec.get("in_time"))
                t_out = self._parse_hhmm(rec.get("out_time"))
                status = rec.get("status") or ""
                if t_in and t_out:
                    minutes = ((t_out.hour - t_in.hour) * 60 + (t_out.minute - t_in.minute)) % (
                        24 * 60
                    )
                    cell.value = round(minutes / 60.0, 2)
                elif status in WORKED_DEFAULT_HOURS:
                    cell.value = WORKED_DEFAULT_HOURS[status]
                elif status:
                    cell.value = status

            rng = f"{first_day}{row}:{last_day}{row}"
            ws[f"AN{row}"] = f"=SUM({rng})"
            ws[f"AO{row}"] = f"=COUNT({rng})"
            ws[f"AP{row}"] = _countif_sum(rng, HOLIDAY_CODES)
            ws[f"AR{row}"] = _countif_sum(rng, LEAVE_CODES)
            ws[f"AS{row}"] = _countif_sum(rng, WEEK_OFF_CODES)
            # Total Paid Days — same semantics the template used.
            ws[f"AT{row}"] = f"=AO{row}+AP{row}+AR{row}+AS{row}"
            ws[f"AN{row}"].number_format = "0.00"

        # Totals + averages directly below the last employee row (the
        # template hardcodes these at rows 13-14 with broken ranges).
        n = len(employees)
        if n:
            total_row = start_row + n
            avg_row = total_row + 1
            last_emp_row = start_row + n - 1
            ws.cell(row=total_row, column=7).value = "Total Working  hours"
            ws.cell(row=avg_row, column=7).value = "Average day worked in this date"
            for c in range(day_min_col, day_max_col + 1):
                col = get_column_letter(c)
                ws[f"{col}{total_row}"] = f"=SUM({col}{start_row}:{col}{last_emp_row})"
                ws[f"{col}{avg_row}"] = f"=ROUND({col}{total_row}/8,2)"
            for col in ("AN", "AO", "AP", "AQ", "AR", "AS", "AT"):
                ws[f"{col}{total_row}"] = f"=SUM({col}{start_row}:{col}{last_emp_row})"

    # ─── Form XXVII — Register of Wages ───────────────────────────────────────

    def _write_form_xxvii(self, writer, header_data, employees, payroll_by_emp):
        sheet_name = "Form XXVII Register of Wages"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 7)

        for idx, emp in enumerate(employees):
            row = start_row + idx
            pr = payroll_by_emp.get(emp.id)

            # Find the salary rule for declared amounts
            # We use the payroll record's earned values + the declared from salary_rules

            data = {
                "serial_no": idx + 1,
                "employee_code": emp.employee_code,
                "employee_name": emp.employee_name,
                "gender": emp.gender,
                "designation": emp.designation,
                "wage_period": "Monthly",
                "days_worked": pr.present_days if pr else 0,
                "national_holidays": pr.holiday_days if pr else 0,
                "ot_hours": pr.ot_hours if pr else 0,
            }

            if pr:
                data.update(
                    {
                        "basic_earned": pr.basic_earned,
                        "da_earned": pr.da_earned,
                        "hra_earned": pr.hra_earned,
                        "med_allowance_earned": pr.medical_allowance,
                        "conv_allowance_earned": pr.conveyance_allowance,
                        "nfh_wages": pr.nfh_wages,
                        "leave_wages": pr.leave_wages,
                        "bonus_earned": pr.bonus_earned,
                        "arrear": pr.arrear,
                    }
                )

            writer.write_row_data(sheet_name, row, data)

    # ─── Form XXVIII — Wage Slip ──────────────────────────────────────────────

    def _write_form_xxviii(self, writer, header_data, employees, payroll_by_emp):
        sheet_name = "Form XXVIII Wage Slip"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        mapping = self.mapper.get_mapping(sheet_name)
        block_size = mapping.get("block_size", 35)
        block_start = mapping.get("block_start_row", 1)

        ws = writer.wb[sheet_name]

        # The template's slip blocks are NOT uniformly sized (35/42/35/37/36
        # rows apart), so locate each block by its "Name of Employee" label
        # and anchor writes to that. Falls back to fixed block_size stepping
        # for employees beyond the pre-built blocks.
        name_offset = mapping.get("columns", {}).get("employee_name", {}).get("row_offset", 9)
        anchor_rows = []
        for r in range(1, min(ws.max_row, 1000) + 1):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, str) and "Name of Employee" in v:
                anchor_rows.append(r - name_offset)

        for idx, emp in enumerate(employees):
            pr = payroll_by_emp.get(emp.id)
            if idx < len(anchor_rows):
                base_row = anchor_rows[idx]
            else:
                logger.warning(
                    f"Wage slip template has only {len(anchor_rows)} blocks; "
                    f"employee {emp.employee_code} written with fixed offset"
                )
                base_row = block_start + (idx * block_size)

            # Write employee details using column offsets
            columns = mapping.get("columns", {})
            for key, ref in columns.items():
                if isinstance(ref, dict):
                    col = ref.get("col", "A")
                    offset = ref.get("row_offset", 0)
                    cell_ref = f"{col}{base_row + offset}"

                    value = None
                    if key == "employee_name":
                        value = emp.employee_name
                    elif key == "father_husband_name":
                        value = emp.father_husband_name or ""
                    elif key == "designation":
                        value = emp.designation
                    elif key == "date_of_joining":
                        value = (
                            emp.date_of_joining.strftime("%d-%m-%Y") if emp.date_of_joining else ""
                        )
                    elif key == "wage_period":
                        value = "Monthly"
                    elif pr:
                        if key == "conv_allowance":
                            value = pr.conveyance_allowance
                        elif key == "med_allowance":
                            value = pr.medical_allowance
                        elif key == "others":
                            value = 0
                        elif key == "others_arrear":
                            value = pr.arrear
                        elif key == "others_deduction":
                            value = pr.other_deductions

                    if value is not None:
                        safe_set(ws, cell_ref, value)

            # Write formula-referenced cells (computed values, not formulas)
            formula_cells = mapping.get("formula_cells", {})
            for key, ref in formula_cells.items():
                if isinstance(ref, dict) and pr:
                    col = ref.get("col", "A")
                    offset = ref.get("row_offset", 0)
                    cell_ref = f"{col}{base_row + offset}"

                    value = None
                    if key == "basic":
                        value = pr.basic_earned
                    elif key == "da":
                        value = pr.da_earned
                    elif key == "hra":
                        value = pr.hra_earned
                    elif key == "nfh_wages":
                        value = pr.nfh_wages
                    elif key == "overtime":
                        value = pr.ot_earned
                    elif key == "gross_wages":
                        value = pr.gross_salary
                    elif key == "pf_deduction":
                        value = pr.pf_deduction
                    elif key == "esi_deduction":
                        value = pr.esi_deduction
                    elif key == "prof_tax":
                        value = pr.professional_tax
                    elif key == "advances":
                        value = pr.advance_deduction
                    elif key == "damages":
                        value = pr.damage_deduction
                    elif key == "lwf":
                        value = pr.lwf
                    elif key == "total_deduction":
                        value = pr.total_deductions
                    elif key == "net_amount":
                        value = pr.net_salary

                    if value is not None:
                        safe_set(ws, cell_ref, value)

    # ─── Form B — Consolidated Wages ──────────────────────────────────────────

    def _write_form_b(self, writer, header_data, employees, payroll_by_emp):
        sheet_name = "Form B Consolidated Wages"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 8)

        # Aggregate totals
        total_emps = len([e for e in employees if payroll_by_emp.get(e.id)])
        total_emoluments = sum(p.gross_salary for p in payroll_by_emp.values())
        total_pf = sum(p.pf_deduction for p in payroll_by_emp.values())
        total_esi = sum(p.esi_deduction for p in payroll_by_emp.values())
        total_advance_fine = sum(
            p.advance_deduction + p.damage_deduction for p in payroll_by_emp.values()
        )
        total_others = sum(p.other_deductions for p in payroll_by_emp.values())
        total_lwf = sum(p.lwf for p in payroll_by_emp.values())

        writer.write_row_data(
            sheet_name,
            start_row,
            {
                "month_date": header_data["month_name"],
                "total_employees": total_emps,
                "total_emoluments": total_emoluments,
                "deductions_epf": total_pf,
                "deductions_esi": total_esi,
                "deductions_advance_fine": total_advance_fine,
                "deductions_others": total_others,
                "deductions_lwf": total_lwf,
            },
        )

    # ─── Form C — Register of Unpaid Accumulations ────────────────────────────

    def _write_form_c(self, writer, header_data):
        """Header-only: unpaid accumulations have no source data model, so the
        register is legitimately empty apart from contractor name and year."""
        sheet_name = "Form C Register of Unpaid"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)

    # ─── Form D — Equal Wages Register ────────────────────────────────────────

    def _write_form_d(self, writer, header_data, employees, payroll_by_emp):
        sheet_name = "Form D Equal Wages Register"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 10)

        # Group by skill category
        categories = {}
        for emp in employees:
            cat = emp.skill_category or "Semi Skilled"
            if cat not in categories:
                categories[cat] = {"men": 0, "women": 0, "total_basic": 0.0, "total_da": 0.0}

            pr = payroll_by_emp.get(emp.id)
            if emp.gender.upper() in ("M", "MALE"):
                categories[cat]["men"] += 1
            else:
                categories[cat]["women"] += 1

            if pr:
                categories[cat]["total_basic"] += pr.basic_earned
                categories[cat]["total_da"] += pr.da_earned

        for idx, (cat, data) in enumerate(categories.items()):
            row = start_row + idx
            total_rate = data["total_basic"] + data["total_da"]
            writer.write_row_data(
                sheet_name,
                row,
                {
                    "category": cat,
                    "number_of_men": data["men"],
                    "number_of_women": data["women"],
                    "total_rate_paid": total_rate,
                    "basic_salary": data["total_basic"],
                    "da": data["total_da"],
                },
            )

        # Write gender totals in header
        total_men = sum(d["men"] for d in categories.values())
        total_women = sum(d["women"] for d in categories.values())
        writer.write_row_data(
            sheet_name,
            4,
            {
                "total_men": total_men,
                "total_women": total_women,
            },
        )

    # ─── Form VI — Holiday Register ───────────────────────────────────────────

    def _write_form_vi(self, writer, header_data, employees):
        sheet_name = "Form VI"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 8)

        for idx, emp in enumerate(employees):
            row = start_row + idx
            writer.write_row_data(
                sheet_name,
                row,
                {
                    "serial_no": idx + 1,
                    "employee_name": emp.employee_name,
                    "father_husband_name": emp.father_husband_name or "",
                    "date_of_joining": emp.date_of_joining.strftime("%d-%m-%Y")
                    if emp.date_of_joining
                    else "",
                    "remarks_left": emp.date_of_leaving.strftime("%d-%m-%Y")
                    if emp.date_of_leaving
                    else "",
                },
            )

    # ─── Deduction Sheet (Form XXIX) ──────────────────────────────────────────

    def _write_deduction_sheet(self, writer, header_data, employees, payroll_by_emp):
        sheet_name = "Deduction"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 7)

        for idx, emp in enumerate(employees):
            row = start_row + idx
            pr = payroll_by_emp.get(emp.id)

            data = {
                "serial_no": idx + 1,
                "employee_code": emp.employee_code,
                "employee_name": emp.employee_name,
                "father_husband_name": emp.father_husband_name or "",
                "designation": emp.designation,
            }

            if pr and (pr.advance_deduction > 0 or pr.damage_deduction > 0):
                data["advance_amount"] = pr.advance_deduction if pr.advance_deduction > 0 else ""
                data["damage_amount"] = pr.damage_deduction if pr.damage_deduction > 0 else ""

            writer.write_row_data(sheet_name, row, data)

    # ─── In-Out Register ──────────────────────────────────────────────────────

    @staticmethod
    def _write_attendance_value(ws, row, col, value, number_format=None, fill_color=None):
        """
        Write one attendance value into its designated cell.

        - Resolves merged cells to their anchor instead of raising.
        - Never overwrites an existing formula (template formulas outside the
          rebuilt band stay intact).
        - Applies the register font (Arial Narrow 28) to the written cell
          only, preserving bold/italic/color; borders, merges, row heights
          and column widths are not touched.
        - Optionally flags the cell red (absent/leave/half-day statuses).
        """
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if cell.coordinate in rng:
                    cell = ws.cell(row=rng.min_row, column=rng.min_col)
                    break
            else:
                logger.warning(f"Merged cell at r{row}c{col} on '{ws.title}' has no range; skipped")
                return None
        existing = cell.value
        is_new_formula = isinstance(value, str) and value.startswith("=")
        if isinstance(existing, str) and existing.startswith("=") and not is_new_formula:
            logger.warning(
                f"Skipping write at {cell.coordinate} on '{ws.title}' — cell holds a formula"
            )
            return None
        cell.value = value
        if number_format:
            cell.number_format = number_format
        old = cell.font
        cell.font = Font(
            name=ATTENDANCE_FONT_NAME,
            size=ATTENDANCE_FONT_SIZE,
            bold=old.bold,
            italic=old.italic,
            color=old.color,
        )
        if fill_color:
            cell.fill = fill_color
        return cell

    def _write_in_out_register(self, writer, header_data, employees, emp_attendance, days_in_month):
        """
        Rebuild the gate-entry register from scratch each generation.

        The template's blocks are non-uniform (a mix of 4-row and 2-row
        employee blocks) and full of stale manual data, so writing into the
        existing grid interleaves new employees with leftovers. Instead the
        whole data band is unmerged, cleared, and rebuilt as uniform 4-row
        blocks: In / Out (swipe times), Actual Hours (live formula from the
        times), OT Hours (live formula, hours beyond 8).
        """
        sheet_name = "In -out register"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 5)
        rows_per_emp = mapping.get("rows_per_employee", 4)
        row_labels = ("In", "Out", "Actual Hours worked", "OT Hours")

        daily = mapping.get("daily_data", {})
        day_min_col = openpyxl.utils.column_index_from_string(daily.get("start_col", "E"))
        max_day_slots = daily.get("max_days", 31)
        day_max_col = day_min_col + max_day_slots - 1  # AI
        total_col, days_col, ot_col = (
            day_max_col + 1,
            day_max_col + 2,
            day_max_col + 3,
        )  # AJ, AK, AL
        first_day = get_column_letter(day_min_col)
        last_day = get_column_letter(day_max_col)

        ws = writer.wb[sheet_name]
        writer.unmerge_data_region(sheet_name, start_row, 1, ot_col)

        # Clear the stale band, keeping the certification footer text.
        certification = None
        stale_end = self._last_used_row(ws, start_row, ot_col, start_row + 300)
        for r in range(start_row, stale_end + 1):
            for c in range(1, ot_col + 1):
                cell = ws.cell(row=r, column=c)
                if (
                    certification is None
                    and isinstance(cell.value, str)
                    and "Certified that" in cell.value
                ):
                    certification = cell.value
                cell.value = None

        # Day headers for this month (template pre-fills only 1..30).
        header_row = start_row - 1
        for slot in range(1, max_day_slots + 1):
            ws.cell(row=header_row, column=day_min_col + slot - 1).value = (
                slot if slot <= days_in_month else None
            )

        # ── Snapshot clean styles from the template's first block BEFORE any
        #    data is written.  Subsequent employee blocks copy from this
        #    snapshot, not from the (already-modified) first block, so fills
        #    applied to employee-1 never bleed into later employees. ──
        _clean_styles: dict[tuple[int, int], object] = {}
        for offset in range(rows_per_emp):
            for c in range(1, ot_col + 1):
                _clean_styles[(offset, c)] = copy(ws.cell(row=start_row + offset, column=c)._style)

        for idx, emp in enumerate(employees):
            r_in = start_row + idx * rows_per_emp
            r_out, r_hours, r_ot = r_in + 1, r_in + 2, r_in + 3

            # Apply the clean template styles to every block (including the
            # first — the stale-band clear above removed values but not
            # fills, so a leftover fill from an old template could survive).
            for offset in range(rows_per_emp):
                for c in range(1, ot_col + 1):
                    ws.cell(row=r_in + offset, column=c)._style = copy(_clean_styles[(offset, c)])

            ws.cell(row=r_in, column=1).value = idx + 1
            ws.cell(row=r_in, column=2).value = emp.employee_code
            ws.cell(row=r_in, column=3).value = emp.employee_name
            for offset, label in enumerate(row_labels):
                ws.cell(row=r_in + offset, column=4).value = label

            daily_records = emp_attendance.get(emp.id, {})
            for day_num in range(1, days_in_month + 1):
                c = day_min_col + day_num - 1
                col = get_column_letter(c)
                rec = daily_records.get(day_num)
                if not rec:
                    continue
                t_in = self._parse_hhmm(rec.get("in_time"))
                t_out = self._parse_hhmm(rec.get("out_time"))
                status = (rec.get("status") or "").upper().strip()
                is_red = status in RED_STATUSES
                if t_in and t_out:
                    fill_color = RED_FILL if is_red else None
                    self._write_attendance_value(ws, r_in, c, t_in, "HH:MM", fill_color=fill_color)
                    self._write_attendance_value(
                        ws, r_out, c, t_out, "HH:MM", fill_color=fill_color
                    )
                    # MOD handles overnight shifts (out < in rolls past midnight)
                    self._write_attendance_value(
                        ws, r_hours, c, f"=ROUND(MOD(({col}{r_out}-{col}{r_in})*24,24),2)", "0.00"
                    )
                    self._write_attendance_value(ws, r_ot, c, f"=MAX(0,{col}{r_hours}-8)", "0.00")
                elif status:
                    # ── Determine the fill colour for this status ──
                    fill_color = None
                    if status in ("W/O", "WO"):
                        fill_color = YELLOW_FILL
                    elif status == "NH":
                        fill_color = BLUE_FILL
                    elif status in RED_STATUSES:
                        fill_color = RED_FILL

                    # Absent / Weekly-Off / National Holiday → merge all 4
                    # rows in this day column and colour the entire block.
                    if status in ("A", "W/O", "WO", "NH"):
                        # Paint all 4 cells BEFORE merging so the fill is
                        # stored on every underlying cell.  Excel renders
                        # the anchor's format across a merge, but writing
                        # all four makes the result robust even if the file
                        # is later un-merged manually.
                        for row_offset in (r_in, r_out, r_hours, r_ot):
                            cell = ws.cell(row=row_offset, column=c)
                            if fill_color:
                                cell.fill = fill_color
                        # Write the status label into the anchor cell.
                        self._write_attendance_value(ws, r_in, c, status, fill_color=fill_color)
                        ws.merge_cells(
                            start_row=r_in,
                            start_column=c,
                            end_row=r_ot,
                            end_column=c,
                        )
                        ws.cell(row=r_in, column=c).alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                        )
                    else:
                        self._write_attendance_value(ws, r_in, c, status, fill_color=fill_color)
                        if status in WORKED_DEFAULT_HOURS:
                            # Marked worked but no swipe times captured yet.
                            self._write_attendance_value(
                                ws,
                                r_hours,
                                c,
                                WORKED_DEFAULT_HOURS[status],
                                "0.00",
                            )
                            self._write_attendance_value(
                                ws,
                                r_ot,
                                c,
                                0,
                                "0.00",
                            )

            hours_rng = f"{first_day}{r_hours}:{last_day}{r_hours}"
            ot_rng = f"{first_day}{r_ot}:{last_day}{r_ot}"
            cell_total = ws.cell(row=r_in, column=total_col)
            cell_total.value = f"=SUM({hours_rng})"
            cell_total.number_format = "0.00"
            cell_days = ws.cell(row=r_in, column=days_col)
            cell_days.value = f"=COUNT({hours_rng})"
            cell_days.number_format = "0"
            cell_ot_total = ws.cell(row=r_in, column=ot_col)
            cell_ot_total.value = f"=SUM({ot_rng})"
            cell_ot_total.number_format = "0.00"

            # Serial/code/name and the three totals span the block, template-style.
            for c in (1, 2, 3, total_col, days_col, ot_col):
                ws.merge_cells(start_row=r_in, start_column=c, end_row=r_ot, end_column=c)

        cert_row = start_row + len(employees) * rows_per_emp + 1
        ws.cell(row=cert_row, column=1).value = certification or DEFAULT_CERTIFICATION

    # ─── ESIC/EPF Tracking ────────────────────────────────────────────────────

    def _write_esic_epf_tracking(self, writer, header_data, employees):
        sheet_name = "ESIC_EPF_Tracking sheet"
        if sheet_name not in writer.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template")
            return

        writer.write_headers(sheet_name, header_data)
        mapping = self.mapper.get_mapping(sheet_name)
        start_row = mapping.get("data_start_row", 14)

        for idx, emp in enumerate(employees):
            row = start_row + idx
            writer.write_row_data(
                sheet_name,
                row,
                {
                    "serial_no": idx + 1,
                    "employee_name": emp.employee_name,
                    "uan": emp.uan or "",
                    "pf_number": emp.pf_number or "",
                    "esic_number": emp.esic_number or "",
                    "location": emp.location or "Coimbatore",
                },
            )

    # ─── Hidden Source Data Sheets ─────────────────────────────────────────────

    def _write_source_data(self, wb, employees, attendance, payroll):
        """
        Write hidden reference sheets containing the source data snapshot.
        Every number in the statutory sheets can be traced back to these.
        """
        # ── Employee Master snapshot ──
        if "_Employee_Master" not in wb.sheetnames:
            ws = wb.create_sheet("_Employee_Master")
        else:
            ws = wb["_Employee_Master"]

        headers = [
            "ID",
            "E-Code",
            "Name",
            "Father/Husband",
            "Gender",
            "DOB",
            "DOJ",
            "DOL",
            "Designation",
            "Skill Category",
            "Status",
            "PF No.",
            "UAN",
            "ESIC No.",
            "Bank A/C",
            "Salary",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, emp in enumerate(employees, 2):
            # str(): WWE OS primary keys are UUIDs, which openpyxl cannot write
            # natively (legacy PKs were ints). Only this audit snapshot sheet
            # emits a raw id — the statutory forms all key off employee_code.
            ws.cell(row=row_idx, column=1, value=str(emp.id))
            ws.cell(row=row_idx, column=2, value=emp.employee_code)
            ws.cell(row=row_idx, column=3, value=emp.employee_name)
            ws.cell(row=row_idx, column=4, value=emp.father_husband_name or "")
            ws.cell(row=row_idx, column=5, value=emp.gender)
            ws.cell(row=row_idx, column=6, value=str(emp.dob) if emp.dob else "")
            ws.cell(row=row_idx, column=7, value=str(emp.date_of_joining))
            ws.cell(
                row=row_idx, column=8, value=str(emp.date_of_leaving) if emp.date_of_leaving else ""
            )
            ws.cell(row=row_idx, column=9, value=emp.designation)
            ws.cell(row=row_idx, column=10, value=emp.skill_category)
            ws.cell(row=row_idx, column=11, value=emp.status)
            ws.cell(row=row_idx, column=12, value=emp.pf_number or "")
            ws.cell(row=row_idx, column=13, value=emp.uan or "")
            ws.cell(row=row_idx, column=14, value=emp.esic_number or "")
            ws.cell(row=row_idx, column=15, value=emp.bank_account or "")
            ws.cell(row=row_idx, column=16, value=emp.salary)

        ws.sheet_state = "hidden"

        # ── Attendance snapshot ──
        if "_Attendance_Snapshot" not in wb.sheetnames:
            ws2 = wb.create_sheet("_Attendance_Snapshot")
        else:
            ws2 = wb["_Attendance_Snapshot"]

        att_headers = ["Emp ID", "Day", "Status", "In Time", "Out Time"]
        for col, header in enumerate(att_headers, 1):
            ws2.cell(row=1, column=col, value=header)

        for row_idx, att in enumerate(attendance, 2):
            # str(): UUID primary key — see the note in the employee sheet above.
            ws2.cell(row=row_idx, column=1, value=str(att.employee_id))
            ws2.cell(row=row_idx, column=2, value=att.day)
            ws2.cell(row=row_idx, column=3, value=att.status)
            ws2.cell(row=row_idx, column=4, value=att.in_time or "")
            ws2.cell(row=row_idx, column=5, value=att.out_time or "")

        ws2.sheet_state = "hidden"
