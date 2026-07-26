import logging
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)


def safe_set(ws, cell_ref: str, value):
    """
    Set a cell value, resolving merged-cell targets to their anchor.

    Templates use merged ranges heavily; openpyxl raises on assignment to a
    non-anchor MergedCell. Writing to the range's top-left anchor is what
    Excel itself does.
    """
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        logger.warning(f"MergedCell {cell_ref} on '{ws.title}' has no containing range; skipped")
        return
    cell.value = value


class ExcelWriter:
    """Writes data to an openpyxl workbook using JSON mappings."""

    def __init__(self, workbook: openpyxl.Workbook, mapper: Any):
        self.wb = workbook
        self.mapper = mapper

    def write_headers(self, sheet_name: str, header_data: dict[str, Any]):
        """Write header values to the specified sheet."""
        try:
            mapping = self.mapper.get_mapping(sheet_name)
            ws = self.wb[sheet_name]
            header_cells = mapping.get("header_cells", {})

            for key, val in header_data.items():
                if key in header_cells:
                    cell_ref = header_cells[key]
                    if isinstance(cell_ref, dict):
                        # Handle complex header refs if needed, simple string for now
                        pass
                    else:
                        safe_set(ws, cell_ref, val)
        except Exception as e:
            logger.error(f"Error writing headers for {sheet_name}: {e}")

    def write_row_data(self, sheet_name: str, row_index: int, row_data: dict[str, Any]):
        """Write a single row of data based on column mappings."""
        try:
            mapping = self.mapper.get_mapping(sheet_name)
            ws = self.wb[sheet_name]
            columns = mapping.get("columns", {})
            start_row = mapping.get("data_start_row")

            # Preserve template row heights for newly populated rows
            if start_row and row_index > start_row and start_row in ws.row_dimensions:
                ws.row_dimensions[row_index].height = ws.row_dimensions[start_row].height

            for key, val in row_data.items():
                if key in columns:
                    col_ref = columns[key]
                    if isinstance(col_ref, dict):
                        # Handle complex col refs
                        col = col_ref.get("col")
                        r_offset = col_ref.get("row_offset", 0)
                        safe_set(ws, f"{col}{row_index + r_offset}", val)
                    else:
                        safe_set(ws, f"{col_ref}{row_index}", val)
        except Exception as e:
            logger.error(f"Error writing row data for {sheet_name}: {e}")

    def unmerge_data_region(self, sheet_name: str, min_row: int, min_col: int, max_col: int):
        """
        Unmerge any merged ranges intersecting a data region (in the working
        copy only — the template file on disk is never modified). Statutory
        templates carry leftover manual merges inside data areas that would
        otherwise swallow per-cell writes.
        """
        try:
            ws = self.wb[sheet_name]
            to_unmerge = [
                r
                for r in list(ws.merged_cells.ranges)
                if r.max_row >= min_row and r.min_col <= max_col and r.max_col >= min_col
            ]
            for r in to_unmerge:
                ws.unmerge_cells(str(r))
            if to_unmerge:
                logger.info(f"Unmerged {len(to_unmerge)} data-area ranges on '{sheet_name}'")
        except Exception as e:
            logger.error(f"Error unmerging data region for {sheet_name}: {e}")
