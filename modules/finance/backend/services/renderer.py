"""Fills the company's own invoice workbook.

`templates/invoice_template.xlsx` is the master format — the real paper
invoice, converted once from the operator's `.xls` (openpyxl reads only the
modern format). It is treated strictly as a *format*: the sample line item and
sample AMC month text that shipped inside it are wiped on every render, and
only layout, styling, merges, column widths and formulas survive.

Two things are worth knowing before changing this file:

* **Row positions are constants, and they are checked.** The template is a
  fixed A4 form, so the item block and totals block sit at known rows. Rather
  than trusting that silently, `_assert_layout` verifies the labels that
  anchor those rows and fails loudly with the cell reference if someone edits
  the workbook and moves them — a wrong-but-plausible invoice is far worse
  than a refused one.

* **Growing the item block is a manual shift.** openpyxl's `insert_rows`
  moves cells but leaves merged ranges, row heights and the print area
  pointing at the old rows, so this module unmerges the affected region first,
  inserts, then re-lays merges and heights at their new positions and stamps
  one canonical style onto every item row. That is why item thirty of a long
  sales invoice comes out with the same borders, fonts and formulas as item
  one.
"""

from __future__ import annotations

import io
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.conf import settings
from finance.backend.models.invoice import TaxMode
from finance.backend.services.computation import ComputedLine, ComputedTotals, tax_labels
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Template geometry ----------------------------------------------------- #
ROW_PERIOD_TEXT = 17  # AMC "For the month of ..." line, above the items
ROW_FIRST_ITEM = 18
# A padding row inside the item block — carries the table's grid borders but
# none of the sample data's formatting. See `_canonical_item_styles`.
ROW_ITEM_GRID = 20
ROW_LAST_ITEM = 26
ROW_TAX_PRIMARY = 27  # SGST, or IGST for an SEZ site
ROW_TAX_SECONDARY = 28  # CGST, blank for an SEZ site
ROW_GROSS = 29  # items + taxes
ROW_ROUND_OFF = 30
ROW_NET_TOTAL = 31
ROW_WORDS = 33
LAST_ROW = 45
ITEM_CAPACITY = ROW_LAST_ITEM - ROW_FIRST_ITEM + 1

COL_SNO = 1  # A
COL_DESC = 2  # B
COL_DESC_END = 6  # F — description is merged B..F
COL_HSN = 7  # G
COL_QTY = 8  # H
COL_UOM = 9  # I
COL_RATE = 10  # J
COL_VALUE = 11  # K
LAST_COL = COL_VALUE

CELL_INVOICE_NUMBER = "F4"
CELL_DATE = "F7"
CELL_CONSIGNEE_NAME = "A12"
CELL_CONSIGNEE_ADDRESS = "A13"
CELL_FACILITY = "F12"

DATE_FORMAT = "%d-%m-%Y"
CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Labels that pin the geometry above. Checked on every render.
_ANCHORS = {
    "A15": "S.No.",
    "F3": "Invoice No.:",
    "A11": "Consignee :",
    "G31": "NET TOTAL :",
    "A32": "( In Words )",
}


class InvoiceTemplateError(RuntimeError):
    """The template workbook is missing or no longer matches the layout this
    renderer fills."""


def template_path() -> Path:
    return Path(settings.INVOICE_TEMPLATE_PATH)


# --- Letterhead ------------------------------------------------------------ #
# The company's own details — name, address, GST number, bank, terms — are
# printed on the template, not configured anywhere. The PDF renderer reads them
# back out of it rather than repeating them in code, so the workbook stays the
# one place they are edited.
CELL_TITLE = "A1"
CELL_COMPANY = "A3"
CELL_ADDRESS = "A4"
CELL_SELLER_GST = "A10"
CELL_BANK_HEADING = "A34"
BANK_ROWS = (35, 36, 37, 38)
CELL_TERMS_HEADING = "A39"
TERMS_ROWS = (41, 42)
CELL_SIGNATURE_FOR = "G41"
CELL_DECLARATION = "A44"
CELL_SIGNATORY = "I45"


@dataclass(frozen=True)
class Letterhead:
    title: str
    company: str
    address: str
    seller_gst: str
    bank_heading: str
    bank: list[tuple[str, str]]
    terms_heading: str
    terms: list[str]
    declaration: str
    signature_for: str
    signatory: str


def _text(worksheet: Worksheet, coordinate: str) -> str:
    value = worksheet[coordinate].value
    return str(value).strip() if value is not None else ""


def read_letterhead() -> Letterhead:
    """Pulls the fixed company block out of the master template."""
    path = template_path()
    if not path.is_file():
        raise InvoiceTemplateError(f"Invoice template not found at {path}.")
    worksheet = openpyxl.load_workbook(path).active
    bank = [
        (_text(worksheet, f"A{row}"), _text(worksheet, f"C{row}"))
        for row in BANK_ROWS
        if _text(worksheet, f"A{row}")
    ]
    return Letterhead(
        title=_text(worksheet, CELL_TITLE),
        company=_text(worksheet, CELL_COMPANY),
        address=_text(worksheet, CELL_ADDRESS),
        seller_gst=_text(worksheet, CELL_SELLER_GST),
        bank_heading=_text(worksheet, CELL_BANK_HEADING),
        bank=bank,
        terms_heading=_text(worksheet, CELL_TERMS_HEADING),
        terms=[_text(worksheet, f"A{row}") for row in TERMS_ROWS if _text(worksheet, f"A{row}")],
        declaration=_text(worksheet, CELL_DECLARATION),
        signature_for=_text(worksheet, CELL_SIGNATURE_FOR),
        signatory=_text(worksheet, CELL_SIGNATORY),
    )


def _assert_layout(worksheet: Worksheet) -> None:
    for coordinate, expected in _ANCHORS.items():
        value = worksheet[coordinate].value
        actual = value.strip() if isinstance(value, str) else value
        if actual != expected.strip():
            raise InvoiceTemplateError(
                f"Invoice template layout changed: expected {expected!r} at {coordinate}, "
                f"found {actual!r}. Update finance/backend/services/renderer.py to match."
            )


def _canonical_item_styles(worksheet: Worksheet) -> list:
    """The style every item row gets, column by column.

    The template's first item row is the one that carried the sample figures,
    so it holds the fonts and number formats an item is meant to print in. The
    blank rows under it are padding and only carry the table's grid. Taking
    the formatting from the first and the borders from a padding row gives one
    style that is right for a filled row *and* keeps the table's vertical
    rules unbroken — which is what makes row 1 and row 30 of a long sales
    invoice look the same.
    """
    styles = []
    for column in range(1, LAST_COL + 1):
        style = copy(worksheet.cell(row=ROW_FIRST_ITEM, column=column)._style)
        style.borderId = worksheet.cell(row=ROW_ITEM_GRID, column=column)._style.borderId
        styles.append(style)
    return styles


def _apply_item_styles(worksheet: Worksheet, styles: list, *, row: int) -> None:
    for column in range(1, LAST_COL + 1):
        worksheet.cell(row=row, column=column)._style = copy(styles[column - 1])


def _grow_item_block(worksheet: Worksheet, *, item_count: int) -> int:
    """Makes room for `item_count` item rows, gives every one of them the same
    styling, and returns how many rows the sheet grew by (0 when the
    template's own capacity was enough)."""
    extra = max(0, item_count - ITEM_CAPACITY)
    item_styles = _canonical_item_styles(worksheet)

    # Everything from the AMC month line downwards is either cleared, rewritten
    # or moved; unmerge it first so neither the insert nor the cell clearing
    # below trips over a merged cell (they are read-only in openpyxl).
    saved = [
        merged.bounds  # (min_col, min_row, max_col, max_row)
        for merged in list(worksheet.merged_cells.ranges)
        if merged.bounds[1] >= ROW_PERIOD_TEXT
    ]
    for min_col, min_row, max_col, max_row in saved:
        worksheet.unmerge_cells(
            start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col
        )

    # Wipe the sample item and sample AMC month text the template shipped with.
    for row in range(ROW_PERIOD_TEXT, ROW_LAST_ITEM + 1):
        for column in range(1, LAST_COL + 1):
            worksheet.cell(row=row, column=column).value = None

    if extra:
        heights = {
            row: worksheet.row_dimensions[row].height
            for row in range(ROW_FIRST_ITEM + 1, LAST_ROW + 1)
        }
        item_row_height = worksheet.row_dimensions[ROW_FIRST_ITEM].height
        worksheet.insert_rows(ROW_FIRST_ITEM + 1, extra)
        for row, height in heights.items():
            worksheet.row_dimensions[row + extra].height = height
        for offset in range(1, extra + 1):
            worksheet.row_dimensions[ROW_FIRST_ITEM + offset].height = item_row_height

    for offset in range(max(item_count, 1)):
        _apply_item_styles(worksheet, item_styles, row=ROW_FIRST_ITEM + offset)

    # Re-lay the merges we took apart. The AMC month line and the item block's
    # own merges are not restored here — `render_invoice` re-creates them, the
    # items uniformly per row so that a 1-item and a 30-item invoice have
    # identically shaped description cells, and the month line only when there
    # is a month to print.
    for min_col, min_row, max_col, max_row in saved:
        if min_row >= ROW_PERIOD_TEXT and max_row <= ROW_LAST_ITEM:
            continue
        if min_row > ROW_FIRST_ITEM:
            min_row += extra
        if max_row > ROW_FIRST_ITEM:
            max_row += extra
        worksheet.merge_cells(
            start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col
        )

    worksheet.print_area = f"A1:{get_column_letter(LAST_COL)}{LAST_ROW + extra}"
    return extra


def _write_items(worksheet: Worksheet, lines: list[ComputedLine]) -> tuple[int, int]:
    """Writes one row per line item, each carrying the template's own
    `value = quantity * rate` formula. Returns the (first, last) item row."""
    first_row = ROW_FIRST_ITEM
    last_row = ROW_FIRST_ITEM + max(len(lines), 1) - 1

    for index, line in enumerate(lines):
        row = first_row + index
        worksheet.merge_cells(
            start_row=row, start_column=COL_DESC, end_row=row, end_column=COL_DESC_END
        )
        worksheet.cell(row=row, column=COL_SNO).value = line.position
        worksheet.cell(row=row, column=COL_DESC).value = line.description
        worksheet.cell(row=row, column=COL_HSN).value = line.hsn
        worksheet.cell(row=row, column=COL_QTY).value = line.quantity
        worksheet.cell(row=row, column=COL_UOM).value = line.uom
        worksheet.cell(row=row, column=COL_RATE).value = line.rate
        worksheet.cell(
            row=row, column=COL_VALUE
        ).value = f"={get_column_letter(COL_QTY)}{row}*{get_column_letter(COL_RATE)}{row}"

    return first_row, last_row


def render_invoice(
    *,
    number: str,
    invoice_date: date,
    consignee_name: str,
    consignee_address: str,
    facility: str,
    lines: list[ComputedLine],
    totals: ComputedTotals,
    tax_mode: str,
    gst_rate: Decimal,
    amount_in_words: str,
    period_text: str = "",
) -> bytes:
    """Returns the filled workbook as bytes. Never writes to disk and never
    mutates the template — storage is the caller's job."""
    path = template_path()
    if not path.is_file():
        raise InvoiceTemplateError(f"Invoice template not found at {path}.")

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    _assert_layout(worksheet)

    extra = _grow_item_block(worksheet, item_count=len(lines))
    first_item_row, last_item_row = _write_items(worksheet, lines)

    value_column = get_column_letter(COL_VALUE)
    items_range = f"{value_column}{first_item_row}:{value_column}{last_item_row}"
    row_tax_primary = ROW_TAX_PRIMARY + extra
    row_tax_secondary = ROW_TAX_SECONDARY + extra
    row_gross = ROW_GROSS + extra
    row_round_off = ROW_ROUND_OFF + extra
    row_net_total = ROW_NET_TOTAL + extra

    # --- Header ----------------------------------------------------------- #
    worksheet[CELL_INVOICE_NUMBER] = number
    worksheet[CELL_DATE] = invoice_date.strftime(DATE_FORMAT)
    worksheet[CELL_CONSIGNEE_NAME] = consignee_name
    worksheet[CELL_CONSIGNEE_ADDRESS] = consignee_address
    worksheet[CELL_FACILITY] = f"Facility: {facility}" if facility else "Facility:"

    # --- AMC month line ---------------------------------------------------- #
    # Sales invoices leave this row blank; it was cleared with the samples.
    if period_text:
        worksheet.merge_cells(
            start_row=ROW_PERIOD_TEXT,
            start_column=COL_DESC,
            end_row=ROW_PERIOD_TEXT,
            end_column=COL_DESC_END,
        )
        worksheet.cell(row=ROW_PERIOD_TEXT, column=COL_DESC).value = period_text

    # --- Tax block --------------------------------------------------------- #
    primary_label, secondary_label = tax_labels(tax_mode=tax_mode, gst_rate=gst_rate)
    rate = Decimal(str(gst_rate))
    primary_rate = rate if tax_mode == TaxMode.IGST else rate / Decimal("2")

    worksheet.cell(row=row_tax_primary, column=COL_RATE).value = primary_label
    worksheet.cell(
        row=row_tax_primary, column=COL_VALUE
    ).value = f"=SUM({items_range})*{_formula_number(primary_rate)}/100"
    if secondary_label:
        worksheet.cell(row=row_tax_secondary, column=COL_RATE).value = secondary_label
        worksheet.cell(
            row=row_tax_secondary, column=COL_VALUE
        ).value = f"=SUM({items_range})*{_formula_number(primary_rate)}/100"
    else:
        worksheet.cell(row=row_tax_secondary, column=COL_RATE).value = None
        worksheet.cell(row=row_tax_secondary, column=COL_VALUE).value = None

    worksheet.cell(
        row=row_gross, column=COL_VALUE
    ).value = f"=SUM({value_column}{first_item_row}:{value_column}{row_tax_secondary})"
    worksheet.cell(row=row_round_off, column=COL_VALUE).value = totals.round_off
    worksheet.cell(
        row=row_net_total, column=COL_VALUE
    ).value = f"=SUM({value_column}{row_gross}:{value_column}{row_round_off})"

    # --- Amount in words --------------------------------------------------- #
    worksheet.cell(row=ROW_WORDS + extra, column=COL_SNO).value = amount_in_words

    return _to_bytes(workbook)


def _formula_number(value: Decimal) -> str:
    return format(Decimal(str(value)).normalize(), "f")


def _to_bytes(workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
