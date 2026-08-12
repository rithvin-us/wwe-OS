"""Invoice-system verification suite.

A single, explicit checklist of the properties the operator's billing depends
on, one test per requirement, exercised through the real generation path
(`InvoiceService.generate`) and the pure engines beneath it. It overlaps the
other finance suites on purpose: this file exists so that "the invoice system
is verified" maps to a named, runnable test for every line of that claim —
AMC, Sales, SEZ, CGST+SGST, IGST, GST maths, numbering, dates, amount in
words, multiple products, template preservation, and file storage.

Nothing here mutates business logic; it only asserts it.
"""

from __future__ import annotations

import hashlib
import io
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from django.conf import settings
from finance.backend.models.invoice import Invoice, InvoiceType, TaxMode
from finance.backend.services.invoice import InvoiceService
from finance.backend.services.numbering import financial_year_for, format_invoice_number
from finance.backend.services.words import rupees_in_words

pytestmark = pytest.mark.django_db

INVOICE_DATE = date(2026, 7, 26)


def _generate(tenant, customer, *, invoice_type=InvoiceType.SALES, lines=None, **extra):
    return InvoiceService().generate(
        tenant=tenant,
        invoice_type=invoice_type,
        invoice_date=INVOICE_DATE,
        customer=customer,
        lines=lines if lines is not None else [{"description": "Service", "rate": 10000}],
        **extra,
    )


def _workbook_values(invoice: Invoice) -> dict:
    """The rendered workbook read back off disk, cell by cell."""
    workbook = openpyxl.load_workbook(Path(invoice.local_path))
    return workbook.active


# --------------------------------------------------------------------------- #
# 1. AMC invoice
# --------------------------------------------------------------------------- #
def test_amc_invoice_generates_with_billed_month(tenant, customer):
    invoice = _generate(tenant, customer, invoice_type=InvoiceType.AMC)
    assert invoice.invoice_type == InvoiceType.AMC
    # Raised 26 Jul 2026 with no explicit period → billed for June (arrears).
    assert (invoice.period_year, invoice.period_month) == (2026, 6)
    assert invoice.period_text == "For the month of June - 2026"
    assert invoice.status == "issued"


# --------------------------------------------------------------------------- #
# 2. Sales invoice
# --------------------------------------------------------------------------- #
def test_sales_invoice_generates_without_month_text(tenant, customer):
    invoice = _generate(tenant, customer, invoice_type=InvoiceType.SALES)
    assert invoice.invoice_type == InvoiceType.SALES
    assert invoice.period_year is None
    assert invoice.period_month is None
    assert invoice.period_text == ""


# --------------------------------------------------------------------------- #
# 3. SEZ invoice
# --------------------------------------------------------------------------- #
def test_sez_invoice_uses_igst_tax_mode(tenant, sez_customer):
    invoice = _generate(tenant, sez_customer)
    assert invoice.is_sez is True
    assert invoice.tax_mode == TaxMode.IGST
    assert invoice.igst_amount > 0
    assert invoice.cgst_amount == Decimal("0.00")
    assert invoice.sgst_amount == Decimal("0.00")


# --------------------------------------------------------------------------- #
# 4. CGST + SGST split
# --------------------------------------------------------------------------- #
def test_cgst_sgst_split_is_half_each(tenant, customer):
    invoice = _generate(tenant, customer)  # subtotal 10,000 @ 18%
    assert invoice.tax_mode == TaxMode.CGST_SGST
    assert invoice.cgst_amount == Decimal("900.00")
    assert invoice.sgst_amount == Decimal("900.00")
    assert invoice.cgst_amount == invoice.sgst_amount
    assert invoice.igst_amount == Decimal("0.00")


# --------------------------------------------------------------------------- #
# 5. IGST full rate
# --------------------------------------------------------------------------- #
def test_igst_is_charged_at_the_full_rate(tenant, sez_customer):
    invoice = _generate(tenant, sez_customer)  # subtotal 10,000 @ 18%
    assert invoice.igst_amount == Decimal("1800.00")
    # IGST equals the two halves a domestic bill would have split into.
    assert invoice.igst_amount == Decimal("900.00") + Decimal("900.00")


# --------------------------------------------------------------------------- #
# 6. GST calculation (rounding, gross, net)
# --------------------------------------------------------------------------- #
def test_gst_calculation_rounds_to_the_nearest_rupee(tenant, customer):
    # 3 × 3333.33 = 9999.99 subtotal; 18% GST; the net rounds to a whole rupee.
    lines = [{"description": "Widget", "quantity": 3, "rate": "3333.33"}]
    invoice = _generate(tenant, customer, lines=lines)
    assert invoice.subtotal == Decimal("9999.99")
    expected_gst = (Decimal("9999.99") * Decimal("18") / Decimal("100")).quantize(Decimal("0.01"))
    assert invoice.cgst_amount + invoice.sgst_amount == expected_gst
    gross = invoice.subtotal + invoice.cgst_amount + invoice.sgst_amount
    assert invoice.total == gross.quantize(Decimal("1"))
    assert (
        invoice.total
        == invoice.subtotal + invoice.cgst_amount + invoice.sgst_amount + invoice.round_off
    )


def test_a_non_default_gst_rate_is_honoured(tenant, customer):
    invoice = _generate(tenant, customer, gst_rate=Decimal("12.00"))  # subtotal 10,000
    assert invoice.gst_rate == Decimal("12.00")
    assert invoice.cgst_amount == Decimal("600.00")
    assert invoice.sgst_amount == Decimal("600.00")


# --------------------------------------------------------------------------- #
# 7. Invoice number generation
# --------------------------------------------------------------------------- #
def test_invoice_numbers_are_sequential_and_formatted(tenant, customer):
    first = _generate(tenant, customer)
    second = _generate(tenant, customer)
    fy = financial_year_for(INVOICE_DATE)
    assert fy == "2026-27"
    assert first.number == format_invoice_number(1, fy) == "G/M/1/2026-27"
    assert second.number == "G/M/2/2026-27"
    assert second.sequence_number == first.sequence_number + 1


def test_amc_and_sales_share_one_number_sequence(tenant, customer):
    amc = _generate(tenant, customer, invoice_type=InvoiceType.AMC)
    sales = _generate(tenant, customer, invoice_type=InvoiceType.SALES)
    assert (amc.sequence_number, sales.sequence_number) == (1, 2)


# --------------------------------------------------------------------------- #
# 8. Date generation
# --------------------------------------------------------------------------- #
def test_invoice_date_is_printed_in_indian_format(tenant, customer):
    invoice = _generate(tenant, customer)
    assert invoice.invoice_date == INVOICE_DATE
    worksheet = _workbook_values(invoice)
    assert worksheet["F7"].value == "26-07-2026"  # dd-mm-yyyy


def test_financial_year_rolls_over_in_april(tenant, customer):
    # A bill dated 1 March 2026 still belongs to FY 2025-26.
    invoice = InvoiceService().generate(
        tenant=tenant,
        invoice_type=InvoiceType.SALES,
        invoice_date=date(2026, 3, 1),
        customer=customer,
        lines=[{"description": "Service", "rate": 10000}],
    )
    assert invoice.financial_year == "2025-26"
    assert invoice.number == "G/M/1/2025-26"


# --------------------------------------------------------------------------- #
# 9. Amount in words
# --------------------------------------------------------------------------- #
def test_amount_in_words_is_stored_and_printed(tenant, customer):
    invoice = _generate(tenant, customer)  # net 11,800
    assert invoice.amount_in_words == rupees_in_words(invoice.total)
    assert invoice.amount_in_words == "Rupees Eleven Thousand Eight Hundred Only"
    worksheet = _workbook_values(invoice)
    assert worksheet["A33"].value == invoice.amount_in_words


@pytest.mark.parametrize(
    "amount,expected",
    [
        (Decimal("11800.00"), "Rupees Eleven Thousand Eight Hundred Only"),
        (
            Decimal("1234567.00"),
            "Rupees Twelve Lakh Thirty Four Thousand Five Hundred Sixty Seven Only",
        ),
        (Decimal("1234.50"), "Rupees One Thousand Two Hundred Thirty Four and Fifty Paise Only"),
        (Decimal("0.00"), "Rupees Zero Only"),
    ],
)
def test_amount_in_words_uses_indian_grouping(amount, expected):
    assert rupees_in_words(amount) == expected


# --------------------------------------------------------------------------- #
# 10. Multiple products
# --------------------------------------------------------------------------- #
def test_multiple_products_are_each_stored_and_summed(tenant, customer):
    lines = [
        {"description": "Pump", "hsn": "8413", "quantity": 2, "rate": "5000.00"},
        {"description": "Valve", "hsn": "8481", "quantity": 3, "rate": "1000.00"},
        {"description": "Gasket", "hsn": "8484", "quantity": 10, "rate": "150.00"},
    ]
    invoice = _generate(tenant, customer, lines=lines)
    stored = list(invoice.lines.order_by("position"))
    assert len(stored) == 3
    assert [line.amount for line in stored] == [
        Decimal("10000.00"),
        Decimal("3000.00"),
        Decimal("1500.00"),
    ]
    assert invoice.subtotal == Decimal("14500.00")
    # Each product lands on its own workbook row from ROW_FIRST_ITEM (18).
    worksheet = _workbook_values(invoice)
    assert worksheet["B18"].value == "Pump"
    assert worksheet["B19"].value == "Valve"
    assert worksheet["B20"].value == "Gasket"


# --------------------------------------------------------------------------- #
# 11. Excel template preservation
# --------------------------------------------------------------------------- #
def test_the_master_template_is_never_mutated(tenant, customer):
    """Generation fills a *copy*; the master workbook on disk must be byte-for
    byte unchanged after any number of renders."""
    template = Path(settings.INVOICE_TEMPLATE_PATH)
    before = hashlib.sha256(template.read_bytes()).hexdigest()
    _generate(
        tenant, customer, lines=[{"description": f"Item {i}", "rate": 100} for i in range(30)]
    )
    after = hashlib.sha256(template.read_bytes()).hexdigest()
    assert before == after


def test_rendered_workbook_keeps_the_letterhead_and_grid(tenant, customer):
    """The fill preserves the template's fixed company block and table anchors,
    proving the master's layout carried through rather than being redrawn."""
    template_ws = openpyxl.load_workbook(settings.INVOICE_TEMPLATE_PATH).active
    invoice = _generate(tenant, customer)
    rendered_ws = _workbook_values(invoice)
    # Anchor labels that pin the layout are identical to the master's.
    for coordinate in ("A11", "A15", "G31", "A32"):
        assert rendered_ws[coordinate].value == template_ws[coordinate].value
    # The company name (letterhead) survives untouched.
    assert rendered_ws["A3"].value == template_ws["A3"].value


# --------------------------------------------------------------------------- #
# 12. Generated file storage
# --------------------------------------------------------------------------- #
def test_generated_files_are_stored_and_downloadable(tenant, customer):
    invoice = _generate(tenant, customer)
    # Both artifacts exist as StoredFile rows.
    assert invoice.file is not None
    assert invoice.pdf_file is not None
    # Filed under the platform's period path, by type.
    assert invoice.file.key.startswith("acme/2026/July/Sales Invoices/")
    assert invoice.file.key.endswith(".xlsx")
    # A local copy exists on disk with real bytes (operator keeps their books).
    xlsx = Path(invoice.local_path)
    assert xlsx.is_file() and xlsx.stat().st_size > 0
    # The stored xlsx opens as a valid workbook.
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx.read_bytes()))
    assert workbook.active is not None
    # The PDF was produced and is a real PDF.
    pdf_path = Path(invoice.pdf_local_path)
    assert pdf_path.is_file()
    assert pdf_path.read_bytes().startswith(b"%PDF")
