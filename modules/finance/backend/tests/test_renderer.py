"""The workbook renderer: the template is a format, the output is a bill.

These tests read the generated workbook back with openpyxl, so they assert on
what actually lands in the file rather than on what the renderer intended.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from finance.backend.models.invoice import TaxMode
from finance.backend.services import computation, renderer


def _render(*, lines, tax_mode=TaxMode.CGST_SGST, period_text="", gst_rate=Decimal("18")):
    computed = computation.compute_lines(lines)
    totals = computation.compute_totals(computed, tax_mode=tax_mode, gst_rate=gst_rate)
    payload = renderer.render_invoice(
        number="G/M/7/2026-27",
        invoice_date=date(2026, 7, 26),
        consignee_name="Waterworks Engineering Pvt Ltd",
        consignee_address="12 Industrial Estate, Coimbatore",
        facility="Peelamedu Plant",
        lines=computed,
        totals=totals,
        tax_mode=tax_mode,
        gst_rate=gst_rate,
        amount_in_words=f"Rupees {totals.total} Only",
        period_text=period_text,
    )
    return openpyxl.load_workbook(io.BytesIO(payload)).active


def _items(count: int, rate=1000):
    return [
        {
            "description": f"Line {index}",
            "hsn": "998719",
            "quantity": 1,
            "uom": "Nos",
            "rate": rate,
        }
        for index in range(1, count + 1)
    ]


# --------------------------------------------------------------------------- #
# Header fields
# --------------------------------------------------------------------------- #
def test_header_fields_are_filled_from_the_invoice():
    sheet = _render(lines=_items(1))
    assert sheet[renderer.CELL_INVOICE_NUMBER].value == "G/M/7/2026-27"
    assert sheet[renderer.CELL_DATE].value == "26-07-2026"
    assert sheet[renderer.CELL_CONSIGNEE_NAME].value == "Waterworks Engineering Pvt Ltd"
    assert sheet[renderer.CELL_CONSIGNEE_ADDRESS].value == "12 Industrial Estate, Coimbatore"
    assert sheet[renderer.CELL_FACILITY].value == "Facility: Peelamedu Plant"


def test_the_templates_sample_placeholders_are_gone():
    sheet = _render(lines=_items(1))
    values = [
        cell.value
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row)
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert not any("<invoice.no>" in value for value in values)
    assert not any("APRIL to JUNE" in value for value in values)


# --------------------------------------------------------------------------- #
# Line items and row expansion
# --------------------------------------------------------------------------- #
def test_one_line_item_lands_on_the_first_item_row():
    sheet = _render(lines=_items(1))
    row = renderer.ROW_FIRST_ITEM
    assert sheet.cell(row=row, column=renderer.COL_SNO).value == 1
    assert sheet.cell(row=row, column=renderer.COL_DESC).value == "Line 1"
    assert sheet.cell(row=row, column=renderer.COL_HSN).value == "998719"
    assert sheet.cell(row=row, column=renderer.COL_UOM).value == "Nos"
    assert sheet.cell(row=row, column=renderer.COL_VALUE).value == f"=H{row}*J{row}"


def test_items_within_the_templates_own_capacity_do_not_move_the_totals():
    sheet = _render(lines=_items(renderer.ITEM_CAPACITY))
    assert sheet.cell(row=renderer.ROW_NET_TOTAL, column=renderer.COL_VALUE).value.startswith(
        "=SUM("
    )
    assert sheet["G31"].value == "NET TOTAL :"


def test_extra_line_items_add_rows_that_carry_the_same_formula():
    count = renderer.ITEM_CAPACITY + 6
    sheet = _render(lines=_items(count))
    first = renderer.ROW_FIRST_ITEM
    for index in range(count):
        row = first + index
        assert sheet.cell(row=row, column=renderer.COL_SNO).value == index + 1
        assert sheet.cell(row=row, column=renderer.COL_VALUE).value == f"=H{row}*J{row}"


def test_every_item_row_is_styled_like_the_first():
    count = renderer.ITEM_CAPACITY + 3
    sheet = _render(lines=_items(count))
    first = sheet.cell(row=renderer.ROW_FIRST_ITEM, column=renderer.COL_VALUE)
    for index in range(1, count):
        row = renderer.ROW_FIRST_ITEM + index
        for column in (renderer.COL_SNO, renderer.COL_DESC, renderer.COL_VALUE):
            source = sheet.cell(row=renderer.ROW_FIRST_ITEM, column=column)
            target = sheet.cell(row=row, column=column)
            assert target.font.name == source.font.name
            assert target.font.sz == source.font.sz
            assert target.number_format == source.number_format
        # The table's vertical rules survive on every row, grown or not.
        assert sheet.cell(row=row, column=renderer.COL_VALUE).border.left.style == "thin"
    assert first.number_format != "General"


def test_growing_the_item_block_pushes_the_totals_block_down_intact():
    extra = 6
    sheet = _render(lines=_items(renderer.ITEM_CAPACITY + extra))
    assert sheet.cell(row=renderer.ROW_NET_TOTAL + extra, column=7).value == "NET TOTAL :"
    assert sheet.cell(row=renderer.ROW_ROUND_OFF + extra, column=8).value == "R/OFF"
    net_total = sheet.cell(row=renderer.ROW_NET_TOTAL + extra, column=renderer.COL_VALUE).value
    assert net_total == f"=SUM(K{renderer.ROW_GROSS + extra}:K{renderer.ROW_ROUND_OFF + extra})"


def test_every_description_cell_is_merged_across_the_description_columns():
    sheet = _render(lines=_items(3))
    merged = {str(area) for area in sheet.merged_cells.ranges}
    for index in range(3):
        row = renderer.ROW_FIRST_ITEM + index
        assert f"B{row}:F{row}" in merged


# --------------------------------------------------------------------------- #
# Tax block
# --------------------------------------------------------------------------- #
def test_a_non_sez_invoice_prints_both_tax_rows():
    sheet = _render(lines=_items(2))
    assert sheet.cell(row=renderer.ROW_TAX_PRIMARY, column=renderer.COL_RATE).value == "SGST 9%"
    assert sheet.cell(row=renderer.ROW_TAX_SECONDARY, column=renderer.COL_RATE).value == "CGST 9%"
    assert sheet.cell(row=renderer.ROW_TAX_PRIMARY, column=renderer.COL_VALUE).value == (
        "=SUM(K18:K19)*9/100"
    )


def test_an_sez_invoice_prints_one_igst_row_and_blanks_the_other():
    sheet = _render(lines=_items(2), tax_mode=TaxMode.IGST)
    assert sheet.cell(row=renderer.ROW_TAX_PRIMARY, column=renderer.COL_RATE).value == "IGST 18%"
    assert sheet.cell(row=renderer.ROW_TAX_SECONDARY, column=renderer.COL_RATE).value is None
    assert sheet.cell(row=renderer.ROW_TAX_PRIMARY, column=renderer.COL_VALUE).value == (
        "=SUM(K18:K19)*18/100"
    )


def test_the_amount_in_words_row_is_filled():
    sheet = _render(lines=_items(1))
    assert str(sheet.cell(row=renderer.ROW_WORDS, column=1).value).startswith("Rupees")


# --------------------------------------------------------------------------- #
# AMC month text
# --------------------------------------------------------------------------- #
def test_an_amc_invoice_prints_the_month_line_above_the_items():
    sheet = _render(lines=_items(1), period_text="For the month of June - 2026")
    assert sheet.cell(row=renderer.ROW_PERIOD_TEXT, column=renderer.COL_DESC).value == (
        "For the month of June - 2026"
    )


def test_a_sales_invoice_leaves_the_month_line_blank():
    sheet = _render(lines=_items(1))
    assert sheet.cell(row=renderer.ROW_PERIOD_TEXT, column=renderer.COL_DESC).value is None


# --------------------------------------------------------------------------- #
# Template integrity
# --------------------------------------------------------------------------- #
def test_a_missing_template_fails_loudly(settings, tmp_path):
    settings.INVOICE_TEMPLATE_PATH = str(tmp_path / "nope.xlsx")
    with pytest.raises(renderer.InvoiceTemplateError):
        _render(lines=_items(1))
